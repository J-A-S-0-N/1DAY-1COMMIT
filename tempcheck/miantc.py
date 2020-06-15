from openpyxl import *
print("todays day")
print("ex : 0410")
da = input(">>")
print("am or pm")
ap = input(">> ")
print("enter your temp")
temp = input(">> ")
wb = load_workbook(filename = "C:\\Users\\user\\Desktop\\G8 오전&오후 체온상황 " + da + ".xlsx")
sheet = wb.active 
if ap == "am":
    c1 = sheet.cell(row = 20, column = 7) 
    c1.value = da
elif ap == "pm":
    c2 = sheet.cell(row= 20 , column = 8) 
    c2.value = da
wb.save(filename = "C:\\Users\\user\\Desktop\\G8 오전&오후 체온상황 " + da + ".xlsx") 